import sys
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure project root is in path
PROJECT_ROOT = Path(r"c:\Users\pgall\Desktop\Python\ChatBot_v4_Grok")
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from chatbot.storage import get_db

def format_date(iso_str):
    if not iso_str:
        return ""
    try:
        # Standardize ISO timestamp to a cleaner date representation
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(iso_str)

def get_leads_report():
    db = get_db()
    leads_col = db["leads"]
    
    query = { "prospecto.codigo": { "$in": ["5783", "67657"] } }
    results = list(leads_col.find(query))
    
    print(f"Encontrados {len(results)} prospectos con la consulta.")
    
    report_data = []
    
    for doc in results:
        prospect = doc.get("prospecto", {})
        messages = doc.get("messages", [])
        
        # Extract first user message
        first_user_msg = ""
        for msg in messages:
            if msg.get("role") == "user":
                first_user_msg = msg.get("content", "").strip()
                break
                
        # Format phone
        phone = doc.get("phone", "")
        if phone and not phone.startswith("+"):
            phone = "+" + phone
            
        row = {
            "Teléfono": phone,
            "Fecha Creación": format_date(doc.get("created_at")),
            "Código Propiedad": prospect.get("codigo", ""),
            "Código MercadoLibre": prospect.get("codigo_mercadolibre", ""),
            "Origen / Portal": prospect.get("origen", ""),
            "Operación": prospect.get("operacion", ""),
            "Tipo": prospect.get("tipo", ""),
            "Comuna": doc.get("comuna") or prospect.get("comuna") or "",
            "Ubicación Referencial": prospect.get("ubicacion_referencial", ""),
            "Ejecutivo Asignado": doc.get("ejecutivo_asignado") or prospect.get("ejecutivo") or "",
            "Estado Pipeline": doc.get("pipeline_stage", ""),
            "Último Intento": doc.get("last_intent", ""),
            "Estado SLA": doc.get("sla_status", ""),
            "Cant. Mensajes": len(messages),
            "Primer Mensaje de Usuario": first_user_msg
        }
        report_data.append(row)
        
    # Create DataFrame
    df = pd.DataFrame(report_data)
    
    # Destination paths
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    if not os.path.exists(desktop_dir):
        desktop_dir = str(PROJECT_ROOT)
        
    output_filename = "Reporte_Leads_5783_67657.xlsx"
    output_path = os.path.join(desktop_dir, output_filename)
    
    # Write to Excel
    print(f"Generando archivo Excel en: {output_path}")
    
    # Use pandas with openpyxl engine
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Leads", index=False)
        
        # Get sheet to apply styles
        workbook = writer.book
        worksheet = writer.sheets["Leads"]
        
        # Styles
        font_family = "Segoe UI"
        
        # Header Styling
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Elegant Dark Blue
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data Cell Styling
        data_font = Font(name=font_family, size=10)
        thin_border_side = Side(border_style="thin", color="D3D3D3")
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Zebra Striping
        zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid") # Soft blue/grey
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Alignments
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # Row heights
        worksheet.row_dimensions[1].height = 28
        
        # Format Headers
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border
            
        # Format Data Rows
        for row_idx in range(2, len(df) + 2):
            worksheet.row_dimensions[row_idx].height = 20
            is_zebra = (row_idx % 2 == 0)
            row_fill = zebra_fill if is_zebra else white_fill
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = cell_border
                
                # Apply alignments based on column type
                col_name = df.columns[col_idx - 1]
                if col_name in ["Teléfono", "Fecha Creación", "Código Propiedad", "Código MercadoLibre", "Operación", "Estado Pipeline", "Estado SLA"]:
                    cell.alignment = align_center
                elif col_name in ["Cant. Mensajes"]:
                    cell.alignment = align_right
                    cell.number_format = "#,##0"
                else:
                    cell.alignment = align_left
                    
        # Autofit Column Widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            col_name = col[0].value
            
            for cell in col:
                # Avoid calculating huge width for the first user message column
                if col_name == "Primer Mensaje de Usuario" and cell.row > 1:
                    max_len = max(max_len, min(len(str(cell.value or "")), 40))
                else:
                    max_len = max(max_len, len(str(cell.value or "")))
                    
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    print("¡Archivo Excel generado exitosamente con estilos premium!")
    return output_path

if __name__ == "__main__":
    try:
        get_leads_report()
    except Exception as e:
        import traceback
        traceback.print_exc()
