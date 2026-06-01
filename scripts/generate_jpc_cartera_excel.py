#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Exporta la cartera de Jorge Pablo Caro y un reporte de informes comunales pendientes."""

from __future__ import annotations

import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(r"C:\Users\pgall\Desktop\Python\ChatBot_v4_Grok")
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from chatbot.lead_router import normalize_text
from chatbot.storage import get_db

OUTPUT_DIR = PROJECT_ROOT / "scripts"
MAIN_FILE = OUTPUT_DIR / "Cartera_Jorge_Pablo_Caro.xlsx"
PENDING_FILE = OUTPUT_DIR / "Pendientes_Informes_Comunales_JPC.xlsx"
ROUND_ROBIN_RM = ["Mariela Arriagada", "Erika Garrido", "Raquel Cheneaux", "Susana Ensignia"]
EXECUTIVE_FILES = {
    "Paula Morales": OUTPUT_DIR / "Cartera_JPC_Paula_Morales.xlsx",
    "Rocío Aliaga": OUTPUT_DIR / "Cartera_JPC_Rocio_Aliaga.xlsx",
    "Mariela Arriagada": OUTPUT_DIR / "Cartera_JPC_Mariela_Arriagada.xlsx",
    "Erika Garrido": OUTPUT_DIR / "Cartera_JPC_Erika_Garrido.xlsx",
    "Raquel Cheneaux": OUTPUT_DIR / "Cartera_JPC_Raquel_Cheneaux.xlsx",
    "Susana Ensignia": OUTPUT_DIR / "Cartera_JPC_Susana_Ensignia.xlsx",
}


def safe_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def first_non_empty(*values: Any) -> str:
    for value in values:
        text = safe_str(value)
        if text:
            return text
    return ""


def normalize_tipo(tipo: Any) -> str:
    txt = normalize_text(tipo or "")
    if "depart" in txt or "depto" in txt:
        return "departamento"
    if "casa" in txt:
        return "casa"
    if "parcela" in txt:
        return "parcela"
    if "terreno" in txt:
        return "terreno"
    if "oficina" in txt:
        return "oficina"
    if "local" in txt:
        return "local comercial"
    return txt or "desconocido"


def detect_region_bucket(region: str, comuna: str) -> str:
    norm_region = normalize_text(region)
    norm_comuna = normalize_text(comuna)

    if any(token in norm_region for token in ("maule", "vii")):
        return "maule"
    if any(token in norm_region for token in ("nuble", "bio", "biobio", "viii", "valparaiso", "quinta", "xvi")):
        return "nuble_biobio_valpo"
    if any(token in norm_region for token in ("metropolitana", "santiago", "xiii")) or norm_comuna in {
        "santiago",
        "providencia",
        "las condes",
        "nunoa",
        "macul",
    }:
        return "rm"
    return "otras"


def route_rm_executive(comuna: str, idx: int) -> str:
    norm_comuna = normalize_text(comuna)
    mariela_priority = {"macul", "nunoa", "providencia", "las condes", "santiago"}

    if norm_comuna in mariela_priority:
        pattern = ["Mariela Arriagada", "Erika Garrido", "Raquel Cheneaux", "Mariela Arriagada", "Erika Garrido", "Susana Ensignia"]
        return pattern[idx % len(pattern)]

    pattern = ["Erika Garrido", "Raquel Cheneaux", "Mariela Arriagada", "Erika Garrido", "Raquel Cheneaux", "Susana Ensignia"]
    return pattern[idx % len(pattern)]


def suggest_executive(region: str, comuna: str, idx: int) -> str:
    bucket = detect_region_bucket(region, comuna)
    if bucket == "maule":
        return "Paula Morales"
    if bucket == "nuble_biobio_valpo":
        return "Rocío Aliaga"
    if bucket == "rm":
        return route_rm_executive(comuna, idx)
    return "Erika Garrido"


def get_market_pairs(db) -> set[tuple[str, str]]:
    pairs = set()
    for doc in db["mercado_comunal"].find({}, {"comuna": 1, "tipo_propiedad": 1}):
        comuna = normalize_text(doc.get("comuna") or "")
        tipo = normalize_text(doc.get("tipo_propiedad") or "")
        if comuna and tipo:
            pairs.add((comuna, tipo))
    return pairs


def get_tasaciones_index(db, codes: list[str]) -> dict[str, Dict[str, Any]]:
    idx: dict[str, Dict[str, Any]] = {}
    for doc in db["tasaciones"].find({"codigo_propiedad": {"$in": codes}}):
        code = safe_str(doc.get("codigo_propiedad"))
        if code:
            idx[code] = doc
    return idx


def build_rows() -> Tuple[pd.DataFrame, pd.DataFrame]:
    db = get_db()
    col = db["universo_cartera"]
    market_pairs = get_market_pairs(db)

    docs = list(
        col.find(
            {
                "disponible": True,
                "oficina": "PROCASA SUCRE",
                "$or": [
                    {"ejecutivo": {"$regex": "jorge pablo caro", "$options": "i"}},
                    {"ejecutivo": {"$regex": "jorge", "$options": "i"}},
                    {"ejecutivo_asignado": {"$regex": "jorge pablo caro", "$options": "i"}},
                ],
            }
        )
    )
    codes = [safe_str(d.get("codigo") or d.get("codigo_propiedad") or d.get("id")) for d in docs]
    tasaciones_idx = get_tasaciones_index(db, [c for c in codes if c])
    rows: List[Dict[str, Any]] = []
    pending_counter: Counter[tuple[str, str]] = Counter()
    pending_examples: Dict[tuple[str, str], Dict[str, Any]] = {}

    for idx, prop in enumerate(sorted(docs, key=lambda d: safe_str(d.get("codigo")))):
        codigo = safe_str(prop.get("codigo") or prop.get("codigo_propiedad") or prop.get("id"))
        nombre = first_non_empty(
            prop.get("nombre_completo_propietario"),
            prop.get("nombre_propietario"),
            prop.get("propietario"),
            prop.get("cliente_nombre"),
        )
        email = first_non_empty(prop.get("email_propietario"), prop.get("correo_propietario"), prop.get("email"), prop.get("correo"))
        fono = first_non_empty(prop.get("fono_propietario"), prop.get("telefono_propietario"), prop.get("telefono"), prop.get("fono"))
        movil = first_non_empty(prop.get("movil_propietario"), prop.get("celular_propietario"), prop.get("movil"))
        tipo = first_non_empty(prop.get("tipo"), prop.get("tipo_propiedad"))
        operacion = first_non_empty(prop.get("operacion"), prop.get("tipo_operacion"))
        region = first_non_empty(prop.get("region"), prop.get("region_nombre"))
        comuna = first_non_empty(prop.get("comuna"), prop.get("comuna_nombre"))
        ejecutivo_original = first_non_empty(prop.get("ejecutivo"), prop.get("ejecutivo_asignado"))

        key = (normalize_text(comuna), normalize_tipo(tipo))
        tas_doc = tasaciones_idx.get(codigo, {})
        tas_status = safe_str(tas_doc.get("status"))
        tas_ok = tas_status == "exito_informe_completo"
        has_comunal = key in market_pairs
        envio = "Tasación" if tas_ok else ("Informe comunal" if has_comunal else "Tasación")
        if key[0] and key[1] and key not in market_pairs:
            pending_counter[key] += 1
            pending_examples.setdefault(key, {"comuna": comuna, "tipo": normalize_tipo(tipo), "codigo_ejemplo": codigo, "region": region})

        sugerido = suggest_executive(region, comuna, idx)
        rows.append(
            {
                "codigo": codigo,
                "nombre completo propietarios": nombre,
                "ejecutivo a cargo (validación)": ejecutivo_original,
                "email propietarios": email,
                "fono propietarios": fono,
                "movil propietarios": movil,
                "tipo propiedad": tipo,
                "operacion": operacion,
                "region": region,
                "comuna": comuna,
                "estado tasacion": tas_status or "sin registro",
                "tiene tasacion exitosa": "SI" if tas_ok else "NO",
                "tiene respaldo comunal": "SI" if has_comunal else "NO",
                "tipo de envío": envio,
                "ejecutivo asignado": sugerido,
                "criterio asignación": detect_region_bucket(region, comuna),
                "fecha export": datetime.now().strftime("%Y-%m-%d %H:%M"),
            }
        )

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(by=["region", "comuna", "tipo propiedad", "codigo"], na_position="last")

    pending_rows = []
    for (comuna_norm, tipo_norm), count in sorted(pending_counter.items(), key=lambda item: (-item[1], item[0][0], item[0][1])):
        sample = pending_examples[(comuna_norm, tipo_norm)]
        pending_rows.append(
            {
                "comuna": sample["comuna"],
                "tipo propiedad": sample["tipo"],
                "cantidad propiedades": count,
                "codigo ejemplo": sample["codigo_ejemplo"],
                "region": sample["region"],
                "accion sugerida": "Descargar informe comunal",
            }
        )

    return df, pd.DataFrame(pending_rows)


def style_excel(path: Path, sheet_name: str) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    zebra_fill = PatternFill("solid", fgColor="F7FBFF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        fill = zebra_fill if row[0].row % 2 == 0 else white_fill
        for cell in row:
            cell.font = data_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 38)
    wb.save(path)


def export_excel(df: pd.DataFrame, path: Path, sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    style_excel(path, sheet_name)


def main() -> None:
    main_df, pending_df = build_rows()
    export_excel(main_df, MAIN_FILE, "Cartera JPC")
    export_excel(pending_df, PENDING_FILE, "Pendientes")

    for executive, path in EXECUTIVE_FILES.items():
        exec_df = main_df[main_df["ejecutivo asignado"] == executive].copy()
        if not exec_df.empty:
            export_excel(exec_df, path, executive[:31])

    print(f"OK: {MAIN_FILE}")
    print(f"OK: {PENDING_FILE}")
    print(f"Filas cartera: {len(main_df)}")
    print(f"Pendientes comunales: {len(pending_df)}")
    print("Distribucion por ejecutivo:")
    for executive in EXECUTIVE_FILES:
        print(f"  - {executive}: {int((main_df['ejecutivo asignado'] == executive).sum())}")


if __name__ == "__main__":
    main()
