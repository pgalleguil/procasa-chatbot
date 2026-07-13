"""
generar_reporte_captacion.py
============================
Genera un reporte Excel completo de captaciones con 6 hojas.
"""

import sys
import os
import io

# Forzar UTF-8 en consola Windows para evitar errores de encoding
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'buffer'):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


import argparse
from datetime import datetime, timezone, timedelta
from collections import defaultdict

# ─── Bootstrap: cargar .env y path del proyecto ────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # Si no hay dotenv, se asume que las variables de entorno ya están seteadas

# Agregar el directorio del proyecto al path para importar config y storage
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
if PROJECT_DIR not in sys.path:
    sys.path.insert(0, PROJECT_DIR)

try:
    from chatbot.storage import get_db
    from chatbot.constants import CHILE_TZ
    from config import Config
except ImportError as e:
    print(f"[ERROR] No se pudo importar módulos del proyecto: {e}")
    print("Asegúrate de ejecutar desde el directorio raíz del proyecto.")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
except ImportError:
    print("[ERROR] openpyxl no está instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

# ─── CONSTANTES DE ESTILO ───────────────────────────────────────────────────
C_NAVY       = "1B2A4A"   # Azul oscuro – encabezados primarios
C_BLUE       = "2563EB"   # Azul medio – encabezados secundarios
C_LIGHT_BLUE = "DBEAFE"   # Azul claro – filas alternadas
C_GREEN      = "16A34A"   # Verde – valores positivos / gestionado
C_GREEN_LIGHT= "DCFCE7"   # Verde claro – fondo positivo
C_AMBER      = "D97706"   # Ámbar – alertas / sin gestionar
C_AMBER_LIGHT= "FEF9C3"   # Ámbar claro
C_RED        = "DC2626"   # Rojo – crítico / descartado
C_RED_LIGHT  = "FEE2E2"   # Rojo claro
C_GRAY       = "F8FAFC"   # Gris muy claro – fondo alternado
C_WHITE      = "FFFFFF"
C_DARK_TEXT  = "111827"
C_MID_TEXT   = "374151"

SLA_DEFECTO  = 5  # días sin gestión = incumple SLA

# ─── HELPERS DE ESTILO ─────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _font(bold=False, color=C_WHITE, size=11, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def _border_thin():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _apply_header_row(ws, row, headers, fill_color=C_NAVY, font_color=C_WHITE, height=30):
    """Escribe una fila de encabezados con estilo completo."""
    ws.row_dimensions[row].height = height
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill   = _fill(fill_color)
        cell.font   = _font(bold=True, color=font_color, size=10)
        cell.alignment = _align("center", "center")
        cell.border = _border_thin()

def _apply_data_row(ws, row, values, alt=False, colors=None):
    """Escribe una fila de datos con estilo alternado."""
    bg = C_LIGHT_BLUE if alt else C_WHITE
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        # Si hay colores específicos por columna, aplicar
        if colors and col - 1 < len(colors) and colors[col - 1]:
            cell.fill = _fill(colors[col - 1])
        else:
            cell.fill = _fill(bg)
        cell.font   = _font(bold=False, color=C_DARK_TEXT, size=10)
        cell.alignment = _align("center", "center")
        cell.border = _border_thin()

def _auto_width(ws, min_w=10, max_w=40):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 4))

def _section_title(ws, row, title, cols=8, fill=C_NAVY):
    """Escribe un título de sección que ocupa varias columnas."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=f"  {title}")
    cell.fill = _fill(fill)
    cell.font = _font(bold=True, size=12, color=C_WHITE)
    cell.alignment = _align("left", "center")
    ws.row_dimensions[row].height = 28

def _kpi_card(ws, row, col, label, value, bg=C_BLUE, fg=C_WHITE):
    """Escribe un mini-KPI (etiqueta + valor) en 2 filas consecutivas."""
    lbl = ws.cell(row=row,   column=col, value=label)
    val = ws.cell(row=row+1, column=col, value=value)
    lbl.fill = _fill(bg); lbl.font = _font(bold=True, size=9, color=fg)
    lbl.alignment = _align("center", "center")
    val.fill = _fill(C_WHITE); val.font = _font(bold=True, size=14, color=bg)
    val.alignment = _align("center", "center")
    ws.row_dimensions[row].height   = 18
    ws.row_dimensions[row+1].height = 30

# ─── EXTRACCIÓN DE DATOS ────────────────────────────────────────────────────

def get_chile_now():
    return datetime.now(CHILE_TZ)

def _ensure_tz(dt):
    """Asegura que un datetime tenga timezone Chile."""
    if dt is None:
        return None
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return None
    if hasattr(dt, 'tzinfo') and dt.tzinfo is None:
        dt = CHILE_TZ.localize(dt)
    elif hasattr(dt, 'tzinfo') and dt.tzinfo:
        dt = dt.astimezone(CHILE_TZ)
    return dt

# Estados que se consideran "gestionados" (al menos un contacto realizado)
ESTADOS_GESTIONADOS = {
    "Contacto exitoso", "Sin respuesta", "Teléfono inválido", "Corredor",
    "Propiedad no disponible", "Publicación expirada", "No interesado",
    "Reunión agendada", "Captado", "Descartado", "Por contactar",
    "GESTION", "CAPTADO", "DESCARTADO"
}
# Estados que se consideran realmente "sin gestión" (no tocadas)
ESTADOS_SIN_GESTION = {"NUEVO", "DETECTADO", None, ""}

# Categorías de resultado para agrupación
GRUPO_POSITIVO   = {"Captado", "CAPTADO", "Reunión agendada", "Contacto exitoso"}
GRUPO_NEUTRO     = {"Sin respuesta", "Por contactar", "GESTION"}
GRUPO_DESCARTADO = {
    "Corredor", "Teléfono inválido", "Descartado", "DESCARTADO",
    "Propiedad no disponible", "Publicación expirada", "No interesado"
}


def cargar_captaciones(db):
    """
    Extrae todas las captaciones relevantes de MongoDB.
    Devuelve una lista de dicts normalizados.
    """
    print("[DATA] Consultando MongoDB...")
    query = {"details.es_propietario_directo": True}
    projection = {
        "_id": 1,
        "url": 1,
        "fecha_captura": 1,
        "fecha": 1,
        "score_captacion": 1,
        "gestion.estado": 1,
        "gestion.estado_captacion": 1,
        "gestion.ejecutivo_asignado": 1,
        "gestion.fecha_ultima_gestion": 1,
        "gestion.fecha_asignacion": 1,
        "gestion.intent_count": 1,
        "gestion.notas": 1,
        "gestion.actividades": 1,
        "details.comuna": 1,
        "details.titulo": 1,
        "details.precio_uf": 1,
        "details.fecha_scraping": 1,
    }

    cursor = db[Config.CAPTACION_COLLECTION_NAME].find(query, projection)
    captaciones = []

    for doc in cursor:
        gestion = doc.get("gestion") or {}
        details = doc.get("details") or {}

        # Fecha de captación (cuándo entró al sistema)
        fecha_raw = (
            details.get("fecha_scraping")
            or doc.get("fecha_captura")
            or doc.get("fecha")
        )
        fecha_captacion = _ensure_tz(fecha_raw)

        # Estado normalizado
        estado = (
            gestion.get("estado_captacion")
            or gestion.get("estado")
            or "NUEVO"
        )
        if estado in ("GESTION", "DETECTADO", "INTENTO DE CONTACTO"):
            estado = "Por contactar"

        # Ejecutivo
        ejecutivo = gestion.get("ejecutivo_asignado") or "Sin asignar"

        # Fecha de última gestión
        fecha_ultima = _ensure_tz(gestion.get("fecha_ultima_gestion"))

        # Notas (historial de gestiones)
        notas_raw = gestion.get("notas") or []
        notas = []
        if isinstance(notas_raw, list):
            for n in notas_raw:
                ts = _ensure_tz(n.get("timestamp"))
                notas.append({
                    "timestamp": ts,
                    "usuario": n.get("usuario", "Sistema"),
                    "canal": n.get("canal", "Manual"),
                    "resultado": n.get("resultado"),
                    "content": n.get("content", ""),
                })

        # Actividades (canal de contacto explícito)
        actividades_raw = gestion.get("actividades") or []
        actividades = []
        if isinstance(actividades_raw, list):
            for a in actividades_raw:
                ts = _ensure_tz(a.get("timestamp"))
                actividades.append({
                    "timestamp": ts,
                    "user": a.get("user", "Sistema"),
                    "canal": a.get("channel") or a.get("canal", "Manual"),
                    "result": a.get("result"),
                })

        # Fecha de primera gestión (primera nota con timestamp)
        timestamps_notas = [n["timestamp"] for n in notas if n["timestamp"]]
        timestamps_acts  = [a["timestamp"] for a in actividades if a["timestamp"]]
        todos_ts = sorted(timestamps_notas + timestamps_acts)
        primera_gestion = todos_ts[0] if todos_ts else None

        # ¿Es gestionada?
        is_gestionada = (
            estado not in ESTADOS_SIN_GESTION
            and (fecha_ultima is not None or len(notas) > 0 or len(actividades) > 0)
        )

        captaciones.append({
            "id": str(doc["_id"]),
            "titulo": details.get("titulo", "Sin título"),
            "comuna": details.get("comuna", "S/I"),
            "precio_uf": details.get("precio_uf"),
            "score": doc.get("score_captacion", 0),
            "estado": estado,
            "ejecutivo": ejecutivo,
            "fecha_captacion": fecha_captacion,
            "fecha_ultima_gestion": fecha_ultima,
            "primera_gestion": primera_gestion,
            "is_gestionada": is_gestionada,
            "intent_count": gestion.get("intent_count", 0),
            "notas": notas,
            "actividades": actividades,
        })

    print(f"[DATA] {len(captaciones)} captaciones cargadas.")
    return captaciones

# ─── HOJA 1: RESUMEN EJECUTIVO ──────────────────────────────────────────────

def hoja_resumen_ejecutivo(wb, captaciones, fecha_reporte):
    ws = wb.create_sheet("1. Resumen Ejecutivo")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4  # margen izquierdo

    # ── Título principal ──
    ws.merge_cells("B1:J1")
    t = ws["B1"]
    t.value = "REPORTE DE CAPTACIONES — RESUMEN EJECUTIVO"
    t.fill  = _fill(C_NAVY)
    t.font  = _font(bold=True, size=16)
    t.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 45

    ws.merge_cells("B2:J2")
    sub = ws["B2"]
    sub.value = f"Generado: {fecha_reporte.strftime('%d/%m/%Y %H:%M')} — Módulo: Captaciones Propietarios Directos"
    sub.fill  = _fill(C_BLUE)
    sub.font  = _font(bold=False, size=10)
    sub.alignment = _align("center", "center")
    ws.row_dimensions[2].height = 20

    # ── KPIs globales ──
    total        = len(captaciones)
    gestionadas  = sum(1 for c in captaciones if c["is_gestionada"])
    sin_gestionar= total - gestionadas
    pct_gestion  = round(gestionadas / total * 100, 1) if total else 0
    captadas     = sum(1 for c in captaciones if c["estado"] in ("Captado", "CAPTADO"))
    pct_captacion= round(captadas / total * 100, 1) if total else 0

    ws.row_dimensions[4].height = 8  # separador

    kpis = [
        ("Total Captaciones", total,       "B", C_NAVY),
        ("Gestionadas",       gestionadas, "D", C_GREEN),
        ("Sin Gestionar",     sin_gestionar,"F", C_AMBER),
        ("% Gestión",         f"{pct_gestion}%", "H", C_BLUE),
        ("Captadas",          captadas,    "J", "7C3AED"),
    ]
    for label, val, col, color in kpis:
        ws.merge_cells(f"{col}5:{col}5")
        ws.merge_cells(f"{col}6:{col}6")
        _kpi_card(ws, 5, ws[f"{col}5"].column, label, val, bg=color)

    ws.row_dimensions[8].height = 8

    # ── Tabla por ejecutivo ──
    _section_title(ws, 9, "📊  PARTICIPACIÓN POR EJECUTIVO", cols=9)

    headers_exec = [
        "Ranking", "Ejecutivo", "Total Captaciones", "% Participación",
        "Gestionadas", "Sin Gestionar", "% Gestión", "Captadas", "% Conversión"
    ]
    _apply_header_row(ws, 10, headers_exec, fill_color=C_NAVY, height=28)

    # Agrupar por ejecutivo
    exec_stats = defaultdict(lambda: {
        "total": 0, "gestionadas": 0, "captadas": 0
    })
    for c in captaciones:
        e = c["ejecutivo"]
        exec_stats[e]["total"] += 1
        if c["is_gestionada"]:
            exec_stats[e]["gestionadas"] += 1
        if c["estado"] in ("Captado", "CAPTADO"):
            exec_stats[e]["captadas"] += 1

    ranking = sorted(exec_stats.items(), key=lambda x: x[1]["total"], reverse=True)

    for i, (exec_name, st) in enumerate(ranking, start=1):
        tot  = st["total"]
        gest = st["gestionadas"]
        sing = tot - gest
        capt = st["captadas"]
        pct_part = round(tot / total * 100, 1) if total else 0
        pct_g    = round(gest / tot * 100, 1) if tot else 0
        pct_conv = round(capt / tot * 100, 1) if tot else 0

        # Color de fondo para % gestión
        if pct_g >= 80:   col_g = C_GREEN_LIGHT
        elif pct_g >= 50: col_g = C_AMBER_LIGHT
        else:             col_g = C_RED_LIGHT

        row_vals  = [i, exec_name, tot, f"{pct_part}%", gest, sing, f"{pct_g}%", capt, f"{pct_conv}%"]
        row_cols  = [None, None, None, None, None, None, col_g, None, None]
        _apply_data_row(ws, 10 + i, row_vals, alt=(i % 2 == 0), colors=row_cols)

    _auto_width(ws)

# ─── HOJA 2: GESTIÓN DE CAPTACIONES ─────────────────────────────────────────

def hoja_gestion_captaciones(wb, captaciones):
    ws = wb.create_sheet("2. Gestión de Captaciones")
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, "📋  GESTIÓN DE CAPTACIONES POR EJECUTIVO", cols=8)

    headers = [
        "Ejecutivo", "Total Asignadas", "Gestionadas", "Sin Gestionar",
        "% Gestión", "Intent Count Total", "Fecha Última Gestión",
        "Días Desde Última Gestión"
    ]
    _apply_header_row(ws, 2, headers, fill_color=C_NAVY)

    now = get_chile_now()

    exec_data = defaultdict(lambda: {
        "total": 0, "gestionadas": 0, "intent_total": 0,
        "ultima_fecha": None, "notas_count": 0
    })
    for c in captaciones:
        e = c["ejecutivo"]
        exec_data[e]["total"] += 1
        if c["is_gestionada"]:
            exec_data[e]["gestionadas"] += 1
        exec_data[e]["intent_total"] += c.get("intent_count", 0)
        fg = c["fecha_ultima_gestion"]
        if fg:
            if exec_data[e]["ultima_fecha"] is None or fg > exec_data[e]["ultima_fecha"]:
                exec_data[e]["ultima_fecha"] = fg

    ranking = sorted(exec_data.items(), key=lambda x: x[1]["total"], reverse=True)
    for i, (exec_name, d) in enumerate(ranking, start=1):
        tot   = d["total"]
        gest  = d["gestionadas"]
        sing  = tot - gest
        pct_g = round(gest / tot * 100, 1) if tot else 0
        uf    = d["ultima_fecha"]
        uf_str = uf.strftime("%d/%m/%Y %H:%M") if uf else "Sin gestiones"
        dias_sin = (now - uf).days if uf else "—"

        if isinstance(dias_sin, int):
            if dias_sin > 7:   col_dias = C_RED_LIGHT
            elif dias_sin > 3: col_dias = C_AMBER_LIGHT
            else:              col_dias = C_GREEN_LIGHT
        else:
            col_dias = C_RED_LIGHT

        cols_color = [None, None, None, None,
                      C_GREEN_LIGHT if pct_g >= 80 else (C_AMBER_LIGHT if pct_g >= 50 else C_RED_LIGHT),
                      None, None, col_dias]
        _apply_data_row(ws, 2 + i, [
            exec_name, tot, gest, sing,
            f"{pct_g}%", d["intent_total"], uf_str, dias_sin
        ], alt=(i % 2 == 0), colors=cols_color)

    # ── Detalle individual de captaciones ──
    row_start = len(ranking) + 5
    _section_title(ws, row_start, "🔍  DETALLE INDIVIDUAL DE CAPTACIONES", cols=8)
    headers_det = [
        "Ejecutivo", "Título", "Estado", "¿Gestionada?",
        "Fecha Captación", "Fecha Últ. Gestión", "Intent Count", "Días Sin Gestión"
    ]
    _apply_header_row(ws, row_start + 1, headers_det, fill_color=C_BLUE)

    for i, c in enumerate(sorted(captaciones, key=lambda x: x["ejecutivo"]), start=1):
        fg = c["fecha_ultima_gestion"]
        fc = c["fecha_captacion"]
        dias_sin = (get_chile_now() - fg).days if fg else "—"
        fc_str = fc.strftime("%d/%m/%Y") if fc else "S/I"
        fg_str = fg.strftime("%d/%m/%Y %H:%M") if fg else "Sin gestión"
        gest_lbl = "✔ Sí" if c["is_gestionada"] else "✘ No"

        bg_gest = C_GREEN_LIGHT if c["is_gestionada"] else C_RED_LIGHT
        cols_color = [None, None, None, bg_gest, None, None, None, None]
        _apply_data_row(ws, row_start + 1 + i, [
            c["ejecutivo"], c["titulo"][:40], c["estado"], gest_lbl,
            fc_str, fg_str, c["intent_count"], dias_sin
        ], alt=(i % 2 == 0), colors=cols_color)

    _auto_width(ws)

# ─── HOJA 3: RESULTADO DE GESTIÓN ───────────────────────────────────────────

def hoja_resultado_gestion(wb, captaciones):
    ws = wb.create_sheet("3. Resultado de Gestión")
    ws.sheet_view.showGridLines = False

    gestionadas = [c for c in captaciones if c["is_gestionada"]]

    # ── Tabla global de estados ──
    _section_title(ws, 1, "📈  RESULTADO DE GESTIÓN — DISTRIBUCIÓN DE ESTADOS", cols=6)
    headers_global = ["Estado / Respuesta", "Cantidad", "% del Total Gestionado",
                      "% del Total General", "Categoría", "Ranking"]
    _apply_header_row(ws, 2, headers_global, fill_color=C_NAVY)

    estado_count = defaultdict(int)
    for c in gestionadas:
        estado_count[c["estado"]] += 1

    total_gest  = len(gestionadas)
    total_gen   = len(captaciones)
    ranking_estados = sorted(estado_count.items(), key=lambda x: x[1], reverse=True)

    for rank, (estado, cnt) in enumerate(ranking_estados, start=1):
        pct_g  = round(cnt / total_gest * 100, 1) if total_gest else 0
        pct_t  = round(cnt / total_gen  * 100, 1) if total_gen  else 0

        if estado in GRUPO_POSITIVO:     cat = "✅ Positivo";    col_cat = C_GREEN_LIGHT
        elif estado in GRUPO_DESCARTADO: cat = "❌ Descartado";  col_cat = C_RED_LIGHT
        else:                            cat = "⏳ En proceso";  col_cat = C_AMBER_LIGHT

        _apply_data_row(ws, 2 + rank, [
            estado, cnt, f"{pct_g}%", f"{pct_t}%", cat, rank
        ], alt=(rank % 2 == 0), colors=[None, None, None, None, col_cat, None])

    # ── Tabla por ejecutivo × estado ──
    row_start = len(ranking_estados) + 5
    _section_title(ws, row_start, "👤  ESTADOS POR EJECUTIVO", cols=6)

    # Calcular ejecutivos únicos y estados únicos
    all_exec   = sorted({c["ejecutivo"] for c in gestionadas})
    all_estados= [e for e, _ in ranking_estados]

    # Encabezado dinámico
    _apply_header_row(ws, row_start + 1, ["Ejecutivo"] + all_estados + ["TOTAL"],
                      fill_color=C_NAVY)

    exec_estado = defaultdict(lambda: defaultdict(int))
    for c in gestionadas:
        exec_estado[c["ejecutivo"]][c["estado"]] += 1

    for i, ex in enumerate(all_exec, start=1):
        row_vals = [ex]
        total_ex = 0
        for est in all_estados:
            v = exec_estado[ex][est]
            row_vals.append(v)
            total_ex += v
        row_vals.append(total_ex)
        _apply_data_row(ws, row_start + 1 + i, row_vals, alt=(i % 2 == 0))

    _auto_width(ws)

# ─── HOJA 4: PRODUCTIVIDAD DIARIA ────────────────────────────────────────────

def hoja_productividad_diaria(wb, captaciones):
    ws = wb.create_sheet("4. Productividad Diaria")
    ws.sheet_view.showGridLines = False

    now = get_chile_now()

    # Recopilar todas las gestiones con fecha
    gestiones = []
    for c in captaciones:
        for nota in c["notas"]:
            if nota["timestamp"]:
                gestiones.append({
                    "fecha": nota["timestamp"].date(),
                    "ejecutivo": nota.get("usuario") or c["ejecutivo"],
                    "canal": nota.get("canal", "Manual"),
                })
        for act in c["actividades"]:
            if act["timestamp"]:
                gestiones.append({
                    "fecha": act["timestamp"].date(),
                    "ejecutivo": act.get("user") or c["ejecutivo"],
                    "canal": act.get("canal", "Manual"),
                })

    # ── Tabla diaria global ──
    _section_title(ws, 1, "📅  GESTIONES POR DÍA — VISIÓN GLOBAL", cols=6)
    headers_dia = ["Fecha", "Día Semana", "Total Gestiones", "Es Día Hábil",
                   "Desviación del Promedio", "Tendencia"]
    _apply_header_row(ws, 2, headers_dia, fill_color=C_NAVY)

    dias_conteo = defaultdict(int)
    for g in gestiones:
        dias_conteo[g["fecha"]] += 1

    DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    dias_sorted = sorted(dias_conteo.keys())

    # Calcular promedio global y semanal
    valores_dia = list(dias_conteo.values())
    prom_global  = round(sum(valores_dia) / len(valores_dia), 1) if valores_dia else 0
    habiles = [v for d, v in dias_conteo.items() if d.weekday() < 5]
    prom_semanal = round(sum(habiles) / len(habiles), 1) if habiles else 0

    dias_cero = [d for d, v in dias_conteo.items() if v == 0]

    for i, fecha in enumerate(dias_sorted, start=1):
        cnt     = dias_conteo[fecha]
        dow     = DIAS_ES[fecha.weekday()]
        es_hab  = "Sí" if fecha.weekday() < 5 else "No"
        desv    = round(cnt - prom_global, 1)
        desv_str= f"+{desv}" if desv >= 0 else str(desv)
        tendencia = "↑" if desv > 0 else ("↓" if desv < 0 else "→")

        col_cnt = C_GREEN_LIGHT if cnt >= prom_global else (C_AMBER_LIGHT if cnt > 0 else C_RED_LIGHT)
        _apply_data_row(ws, 2 + i, [
            fecha.strftime("%d/%m/%Y"), dow, cnt, es_hab, desv_str, tendencia
        ], alt=(i % 2 == 0), colors=[None, None, col_cnt, None, None, None])

    # ── KPIs de productividad ──
    row_kpi = len(dias_sorted) + 5
    _section_title(ws, row_kpi, "⚡  MÉTRICAS DE PRODUCTIVIDAD", cols=6)

    kpis = [
        ("Promedio Diario (Global)",    prom_global,  "B"),
        ("Promedio Días Hábiles",       prom_semanal, "D"),
        ("Días con Cero Gestión",       len(dias_cero),"F"),
        ("Total Gestiones Registradas", len(gestiones),"H"),
    ]
    for label, val, col_l in kpis:
        c_idx = openpyxl.utils.column_index_from_string(col_l)
        _kpi_card(ws, row_kpi + 1, c_idx, label, val, bg=C_NAVY)

    # ── Tabla ejecutivo × día ──
    row_exec = row_kpi + 5
    _section_title(ws, row_exec, "👤  GESTIONES POR EJECUTIVO POR DÍA", cols=6)

    exec_dia = defaultdict(lambda: defaultdict(int))
    for g in gestiones:
        exec_dia[g["ejecutivo"]][g["fecha"]] += 1

    all_exec_prod = sorted(exec_dia.keys())
    # Solo últimos 30 días para no colapsar la hoja
    ultimos_30 = [d for d in dias_sorted if (now.date() - d).days <= 30]
    _apply_header_row(ws, row_exec + 1,
                      ["Ejecutivo"] + [d.strftime("%d/%m") for d in ultimos_30] + ["TOTAL"],
                      fill_color=C_BLUE)

    for i, ex in enumerate(all_exec_prod, start=1):
        row_vals = [ex]
        tot_ex = 0
        for d in ultimos_30:
            v = exec_dia[ex][d]
            row_vals.append(v)
            tot_ex += v
        row_vals.append(tot_ex)
        _apply_data_row(ws, row_exec + 1 + i, row_vals, alt=(i % 2 == 0))

    _auto_width(ws)

# ─── HOJA 5: SEGUIMIENTO COMERCIAL ──────────────────────────────────────────

def hoja_seguimiento_comercial(wb, captaciones, sla_dias=SLA_DEFECTO):
    ws = wb.create_sheet("5. Seguimiento Comercial")
    ws.sheet_view.showGridLines = False

    now = get_chile_now()

    # ── Captaciones nuevas por día ──
    _section_title(ws, 1, f"🔎  SEGUIMIENTO COMERCIAL  |  SLA: {sla_dias} días", cols=7)

    capt_por_dia = defaultdict(int)
    for c in captaciones:
        fc = c["fecha_captacion"]
        if fc:
            capt_por_dia[fc.date()] += 1

    headers_nd = ["Fecha", "Día Semana", "Captaciones Nuevas", "Acumulado Semana"]
    _apply_header_row(ws, 2, headers_nd, fill_color=C_NAVY)
    DIAS_ES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    dias_sorted = sorted(capt_por_dia.keys())
    acum_sem = 0
    semana_actual = None
    for i, fecha in enumerate(dias_sorted, start=1):
        iso_week = fecha.isocalendar()[1]
        if iso_week != semana_actual:
            semana_actual = iso_week
            acum_sem = 0
        acum_sem += capt_por_dia[fecha]
        dow = DIAS_ES[fecha.weekday()]
        _apply_data_row(ws, 2 + i, [
            fecha.strftime("%d/%m/%Y"), dow, capt_por_dia[fecha], acum_sem
        ], alt=(i % 2 == 0))

    # ── Tiempos de respuesta ──
    row_tr = len(dias_sorted) + 5
    _section_title(ws, row_tr, "⏱️  TIEMPO ENTRE CAPTACIÓN Y PRIMERA GESTIÓN", cols=7)
    headers_tr = ["Ejecutivo", "Título", "Fecha Captación", "Primera Gestión",
                  "Días hasta 1ª Gestión", "¿Dentro del SLA?", "Estado"]
    _apply_header_row(ws, row_tr + 1, headers_tr, fill_color=C_NAVY)

    tiempos = []
    for c in captaciones:
        if c["fecha_captacion"] and c["primera_gestion"]:
            delta = (c["primera_gestion"] - c["fecha_captacion"]).total_seconds() / 86400
            tiempos.append((c, round(delta, 1)))

    tiempos_sorted = sorted(tiempos, key=lambda x: x[1])

    for i, (c, delta) in enumerate(tiempos_sorted, start=1):
        dentro_sla = delta <= sla_dias
        sla_str = f"✅ Sí ({delta}d)" if dentro_sla else f"❌ No ({delta}d)"
        col_sla = C_GREEN_LIGHT if dentro_sla else C_RED_LIGHT
        fc_str  = c["fecha_captacion"].strftime("%d/%m/%Y") if c["fecha_captacion"] else "S/I"
        pg_str  = c["primera_gestion"].strftime("%d/%m/%Y") if c["primera_gestion"] else "S/I"
        _apply_data_row(ws, row_tr + 1 + i, [
            c["ejecutivo"], c["titulo"][:35], fc_str, pg_str,
            delta, sla_str, c["estado"]
        ], alt=(i % 2 == 0), colors=[None, None, None, None, None, col_sla, None])

    # ── Pendientes de gestionar y con >X días sin gestión ──
    row_pend = row_tr + len(tiempos_sorted) + 4
    _section_title(ws, row_pend, f"🚨  CAPTACIONES PENDIENTES Y CON MÁS DE {sla_dias} DÍAS SIN GESTIÓN", cols=7)
    headers_pend = ["Ejecutivo", "Título", "Estado", "Fecha Captación",
                    "Días Sin Gestión", "Prioridad", "Score"]
    _apply_header_row(ws, row_pend + 1, headers_pend, fill_color=C_RED)

    pendientes = []
    for c in captaciones:
        if not c["is_gestionada"]:
            fc = c["fecha_captacion"]
            dias_sin = (now - fc).days if fc else 9999
            pendientes.append((c, dias_sin))
        elif c["fecha_ultima_gestion"]:
            dias_sin = (now - c["fecha_ultima_gestion"]).days
            if dias_sin > sla_dias:
                pendientes.append((c, dias_sin))

    pendientes_sorted = sorted(pendientes, key=lambda x: x[1], reverse=True)

    for i, (c, dias_sin) in enumerate(pendientes_sorted, start=1):
        fc_str = c["fecha_captacion"].strftime("%d/%m/%Y") if c["fecha_captacion"] else "S/I"
        if dias_sin > 14:   prioridad = "🔴 CRÍTICA"
        elif dias_sin > 7:  prioridad = "🟡 ALTA"
        else:               prioridad = "🟢 MEDIA"
        col_p = C_RED_LIGHT if dias_sin > 14 else (C_AMBER_LIGHT if dias_sin > 7 else C_GREEN_LIGHT)
        _apply_data_row(ws, row_pend + 1 + i, [
            c["ejecutivo"], c["titulo"][:35], c["estado"], fc_str,
            dias_sin, prioridad, c["score"]
        ], alt=(i % 2 == 0), colors=[None, None, None, None, col_p, col_p, None])

    # ── KPIs finales ──
    row_kpi = row_pend + len(pendientes_sorted) + 4
    _section_title(ws, row_kpi, "📊  RESUMEN DE SEGUIMIENTO", cols=7)

    total          = len(captaciones)
    pend_tot       = sum(1 for c in captaciones if not c["is_gestionada"])
    sla_ok_count   = sum(1 for _, d in tiempos if d <= sla_dias)
    sla_fail_count = len(tiempos) - sla_ok_count
    pct_sla        = round(sla_ok_count / len(tiempos) * 100, 1) if tiempos else 0

    kpis_seg = [
        ("Total Captaciones",     total,         "B"),
        ("Pendientes Gestionar",  pend_tot,       "D"),
        ("Gestionadas Dentro SLA",sla_ok_count,   "F"),
        ("% Cumplimiento SLA",    f"{pct_sla}%",  "H"),
    ]
    for label, val, col_l in kpis_seg:
        c_idx = openpyxl.utils.column_index_from_string(col_l)
        _kpi_card(ws, row_kpi + 1, c_idx, label, val, bg=C_NAVY)

    _auto_width(ws)

# ─── HOJA 6: DASHBOARD ──────────────────────────────────────────────────────

def hoja_dashboard(wb, captaciones):
    ws = wb.create_sheet("6. Dashboard")
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, "📊  DASHBOARD VISUAL — CAPTACIONES", cols=12)

    now = get_chile_now()
    total = len(captaciones)

    # ── Datos para gráficos: captaciones por ejecutivo ──
    exec_stats = defaultdict(lambda: {"total": 0, "gestionadas": 0, "captadas": 0})
    for c in captaciones:
        e = c["ejecutivo"]
        exec_stats[e]["total"] += 1
        if c["is_gestionada"]:
            exec_stats[e]["gestionadas"] += 1
        if c["estado"] in ("Captado", "CAPTADO"):
            exec_stats[e]["captadas"] += 1

    ranking_exec = sorted(exec_stats.items(), key=lambda x: x[1]["total"], reverse=True)

    # ── Escribir tabla de datos para Gráfico 1 (Captaciones por ejecutivo) ──
    ws.cell(row=3, column=1, value="EJECUTIVO").font = _font(bold=True, color=C_DARK_TEXT, size=10)
    ws.cell(row=3, column=2, value="CAPTACIONES").font = _font(bold=True, color=C_DARK_TEXT, size=10)
    ws.cell(row=3, column=3, value="GESTIONADAS").font = _font(bold=True, color=C_DARK_TEXT, size=10)
    ws.cell(row=3, column=4, value="% GESTIÓN").font = _font(bold=True, color=C_DARK_TEXT, size=10)
    ws.cell(row=3, column=5, value="CAPTADAS").font = _font(bold=True, color=C_DARK_TEXT, size=10)

    for i, (exec_name, st) in enumerate(ranking_exec, start=1):
        tot  = st["total"]
        gest = st["gestionadas"]
        capt = st["captadas"]
        pct_g = round(gest / tot * 100, 1) if tot else 0
        ws.cell(row=3+i, column=1, value=exec_name)
        ws.cell(row=3+i, column=2, value=tot)
        ws.cell(row=3+i, column=3, value=gest)
        ws.cell(row=3+i, column=4, value=pct_g)
        ws.cell(row=3+i, column=5, value=capt)

    n_exec = len(ranking_exec)
    data_end_row = 3 + n_exec

    # ── Gráfico 1: Captaciones y Gestiones por Ejecutivo (Barras apiladas) ──
    chart1 = BarChart()
    chart1.type    = "col"
    chart1.grouping = "clustered"
    chart1.title  = "Captaciones vs Gestiones por Ejecutivo"
    chart1.style  = 10
    chart1.y_axis.title = "Cantidad"
    chart1.x_axis.title = "Ejecutivo"
    chart1.width  = 22
    chart1.height = 14

    cats = Reference(ws, min_col=1, min_row=4, max_row=data_end_row)
    data_capt = Reference(ws, min_col=2, min_row=3, max_row=data_end_row)
    data_gest = Reference(ws, min_col=3, min_row=3, max_row=data_end_row)

    chart1.add_data(data_capt, titles_from_data=True)
    chart1.add_data(data_gest, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.series[0].graphicalProperties.solidFill = C_NAVY
    chart1.series[1].graphicalProperties.solidFill = C_GREEN

    ws.add_chart(chart1, "G3")

    # ── Gráfico 2: % de Gestión por Ejecutivo (Barras horizontales) ──
    chart2 = BarChart()
    chart2.type     = "bar"
    chart2.grouping = "clustered"
    chart2.title    = "% Gestión por Ejecutivo"
    chart2.style    = 10
    chart2.y_axis.title = "% Gestión"
    chart2.width    = 22
    chart2.height   = 14

    data_pct = Reference(ws, min_col=4, min_row=3, max_row=data_end_row)
    chart2.add_data(data_pct, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.series[0].graphicalProperties.solidFill = C_BLUE

    ws.add_chart(chart2, "G21")

    # ── Datos para Gráfico 3: Distribución de estados ──
    estado_count = defaultdict(int)
    for c in captaciones:
        estado_count[c["estado"]] += 1

    row_est = data_end_row + 3
    ws.cell(row=row_est, column=1, value="ESTADO").font = _font(bold=True, color=C_DARK_TEXT, size=10)
    ws.cell(row=row_est, column=2, value="CANTIDAD").font = _font(bold=True, color=C_DARK_TEXT, size=10)

    estados_sorted = sorted(estado_count.items(), key=lambda x: x[1], reverse=True)
    for i, (est, cnt) in enumerate(estados_sorted, start=1):
        ws.cell(row=row_est+i, column=1, value=est)
        ws.cell(row=row_est+i, column=2, value=cnt)

    n_est = len(estados_sorted)
    est_end = row_est + n_est

    # ── Gráfico 3: Distribución de Respuestas / Estados (Pie) ──
    chart3 = PieChart()
    chart3.title  = "Distribución de Respuestas (Estados)"
    chart3.style  = 10
    chart3.width  = 22
    chart3.height = 14

    labels3  = Reference(ws, min_col=1, min_row=row_est+1, max_row=est_end)
    data3    = Reference(ws, min_col=2, min_row=row_est,   max_row=est_end)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(labels3)
    chart3.dataLabels = DataLabelList()
    chart3.dataLabels.showPercent = True
    chart3.dataLabels.showCatName = False

    ws.add_chart(chart3, "A38")

    # ── Datos para Gráfico 4: Evolución diaria (últimos 30 días) ──
    gestiones_dia = defaultdict(int)
    for c in captaciones:
        for nota in c["notas"]:
            if nota["timestamp"]:
                d = nota["timestamp"].date()
                if (now.date() - d).days <= 30:
                    gestiones_dia[d] += 1
        for act in c["actividades"]:
            if act["timestamp"]:
                d = act["timestamp"].date()
                if (now.date() - d).days <= 30:
                    gestiones_dia[d] += 1

    row_ev = est_end + 3
    ws.cell(row=row_ev, column=1, value="FECHA").font = _font(bold=True, color=C_DARK_TEXT, size=10)
    ws.cell(row=row_ev, column=2, value="GESTIONES").font = _font(bold=True, color=C_DARK_TEXT, size=10)

    dias_evol = sorted(gestiones_dia.keys())
    for i, d in enumerate(dias_evol, start=1):
        ws.cell(row=row_ev+i, column=1, value=d.strftime("%d/%m"))
        ws.cell(row=row_ev+i, column=2, value=gestiones_dia[d])

    n_ev = len(dias_evol)
    ev_end = row_ev + n_ev

    if n_ev > 0:
        chart4 = LineChart()
        chart4.title    = "Evolución Diaria de Gestiones (últimos 30 días)"
        chart4.style    = 10
        chart4.y_axis.title = "Gestiones"
        chart4.x_axis.title = "Fecha"
        chart4.width    = 22
        chart4.height   = 14
        chart4.grouping = "standard"

        data_ev = Reference(ws, min_col=2, min_row=row_ev, max_row=ev_end)
        cats_ev = Reference(ws, min_col=1, min_row=row_ev+1, max_row=ev_end)
        chart4.add_data(data_ev, titles_from_data=True)
        chart4.set_categories(cats_ev)
        chart4.series[0].graphicalProperties.line.solidFill = C_BLUE
        chart4.series[0].smooth = True

        ws.add_chart(chart4, "G38")

    _auto_width(ws)

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generador de Reporte Excel de Captaciones")
    parser.add_argument("--sla",    type=int,   default=SLA_DEFECTO,
                        help=f"Días SLA para gestión (default: {SLA_DEFECTO})")
    parser.add_argument("--output", type=str,   default="",
                        help="Nombre del archivo de salida (default: reporte_captaciones_FECHA.xlsx)")
    args = parser.parse_args()

    fecha_reporte = datetime.now(CHILE_TZ)
    fecha_str     = fecha_reporte.strftime("%Y%m%d_%H%M")

    if not args.output:
        exports_dir = os.path.join(PROJECT_DIR, "exports")
        os.makedirs(exports_dir, exist_ok=True)
        output_path = os.path.join(exports_dir, f"reporte_captaciones_{fecha_str}.xlsx")
    else:
        output_path = args.output

    print("=" * 60)
    print("  GENERADOR DE REPORTE DE CAPTACIONES")
    print(f"  SLA configurado: {args.sla} días")
    print(f"  Salida: {output_path}")
    print("=" * 60)

    # Conexión a BD
    try:
        db = get_db()
        db.list_collection_names()  # Test de conexión
        print("[DB] Conexión exitosa a MongoDB.")
    except Exception as e:
        print(f"[ERROR] No se pudo conectar a MongoDB: {e}")
        sys.exit(1)

    # Carga de datos
    captaciones = cargar_captaciones(db)
    if not captaciones:
        print("[WARN] No se encontraron captaciones. El reporte se generará vacío.")

    # Generar workbook
    wb = Workbook()
    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("[EXCEL] Generando hoja 1: Resumen Ejecutivo...")
    hoja_resumen_ejecutivo(wb, captaciones, fecha_reporte)

    print("[EXCEL] Generando hoja 2: Gestión de Captaciones...")
    hoja_gestion_captaciones(wb, captaciones)

    print("[EXCEL] Generando hoja 3: Resultado de Gestión...")
    hoja_resultado_gestion(wb, captaciones)

    print("[EXCEL] Generando hoja 4: Productividad Diaria...")
    hoja_productividad_diaria(wb, captaciones)

    print("[EXCEL] Generando hoja 5: Seguimiento Comercial...")
    hoja_seguimiento_comercial(wb, captaciones, sla_dias=args.sla)

    print("[EXCEL] Generando hoja 6: Dashboard...")
    hoja_dashboard(wb, captaciones)

    # Guardar
    try:
        wb.save(output_path)
        gestionadas = sum(1 for c in captaciones if c["is_gestionada"])
        ejecutivos  = {c['ejecutivo'] for c in captaciones}
        print("\n[OK] Reporte generado exitosamente:")
        print(f"     {output_path}")
        print("\n[RESUMEN]")
        print(f"  Total captaciones procesadas : {len(captaciones)}")
        print(f"  Gestionadas                  : {gestionadas}")
        print(f"  Sin gestionar                : {len(captaciones) - gestionadas}")
        print(f"  Ejecutivos encontrados       : {len(ejecutivos)}")
    except Exception as e:
        print(f"[ERROR] No se pudo guardar el archivo: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
