#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Motor operacional de inteligencia comercial.

Construye la coleccion final `propiedades_accionables` usando:
- universo_cartera (base universal)
- tasaciones (si existe tasacion individual)
- mercado_comunal (fallback y señales de mercado)

Salida adicional:
- CSV operacional: exports/propiedades_accionables.csv
"""

from __future__ import annotations

import argparse
import csv
import logging
import math
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pymongo import ASCENDING, DESCENDING, UpdateOne
from pymongo.collection import Collection
from tqdm import tqdm

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from chatbot.storage import get_db
from scripts.comercial_normalization import build_match_key, normalize_comuna_tipo, normalize_tipo_propiedad

logger = logging.getLogger("inteligencia_comercial_pipeline")


def safe_float(v: Any) -> float:
    try:
        x = float(v)
        if math.isnan(x) or math.isinf(x):
            return 0.0
        return x
    except Exception:
        return 0.0


def safe_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in {"1", "true", "si", "sí", "yes"}
    return bool(v)


def clamp_score(score: float) -> float:
    return round(max(0.0, min(100.0, score)), 2)


def parse_dt(value: Any) -> Optional[datetime]:
    if not value:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def ensure_indexes(col: Collection) -> None:
    col.create_index([("codigo_propiedad", ASCENDING)], unique=True, name="uq_codigo_propiedad")
    col.create_index([("score_comercial", DESCENDING)], name="ix_score_comercial")
    col.create_index([("campana_recomendada", ASCENDING)], name="ix_campana_recomendada")
    col.create_index([("riesgo_comercial", ASCENDING)], name="ix_riesgo_comercial")
    col.create_index([("comuna", ASCENDING), ("tipo_propiedad", ASCENDING)], name="ix_comuna_tipo")
    col.create_index([("ready_para_campana", ASCENDING)], name="ix_ready_para_campana")


def infer_tasacion_uf_from_comuna(
    precio_publicado_uf: float,
    m2_construidos: float,
    uf_m2_venta_efectiva_actual: float,
) -> Tuple[float, str]:
    if uf_m2_venta_efectiva_actual > 0 and m2_construidos > 0:
        return round(uf_m2_venta_efectiva_actual * m2_construidos, 2), "uf_m2_x_m2"

    # Sin m2, no inferimos tasacion numérica para evitar inventar valorizacion puntual.
    return 0.0, "senales_comunales"


def extract_tasacion_uf(tas_doc: Optional[Dict[str, Any]], operacion: str) -> Tuple[float, str]:
    if not tas_doc:
        return 0.0, "sin_tasacion"

    tas = tas_doc.get("tasacion_online") or {}
    vcom = tas.get("valor_comercial") or {}
    varr = tas.get("arriendo_estimado") or {}
    vmm = tas.get("valor_minimo_maximo") or {}

    op = (operacion or "").strip().lower()
    if op == "arriendo":
        candidates = [
            ("tasacion_online.arriendo_estimado.uf", safe_float(varr.get("uf"))),
            ("tasacion_online.valor_comercial.uf", safe_float(vcom.get("uf"))),
        ]
    else:
        candidates = [
            ("tasacion_online.valor_comercial.uf", safe_float(vcom.get("uf"))),
            ("tasacion_online.valor_minimo_maximo.estimacion_valor_comercial_uf", safe_float(vmm.get("estimacion_valor_comercial_uf"))),
            ("tasacion_online.valor_minimo_maximo.precio_maximo_uf", safe_float(vmm.get("precio_maximo_uf"))),
            ("tasacion_online.valor_minimo_maximo.precio_minimo_uf", safe_float(vmm.get("precio_minimo_uf"))),
        ]
    for fuente, val in candidates:
        if val > 0:
            return val, fuente
    return 0.0, "tasacion_con_valor_uf_cero"


def risk_level(score: float) -> str:
    if score >= 85:
        return "critico"
    if score >= 70:
        return "alto"
    if score >= 45:
        return "medio"
    return "bajo"


def is_revision_datos(comuna: str, tipo: str, precio_publicado_uf: float) -> bool:
    if not comuna.strip() or comuna.strip().lower() == "desconocido":
        return True
    if not tipo.strip() or tipo.strip().lower() == "desconocido":
        return True
    if precio_publicado_uf <= 0:
        return True
    return False


def build_motivos(
    sobreprecio_pct: float,
    liquidez: str,
    competencia: str,
    brecha_pct: float,
    score_presion: float,
    estado_precio_tasacion: str,
    publicaciones_activas: int,
    publicaciones_totales: int,
    revision_datos_flag: bool,
) -> List[str]:
    motivos: List[str] = []

    if revision_datos_flag:
        motivos.append("revision_datos_faltantes")

    if sobreprecio_pct >= 20:
        motivos.append("sobreprecio_alto")
    elif sobreprecio_pct >= 10:
        motivos.append("sobreprecio_moderado")
    elif sobreprecio_pct >= 5:
        motivos.append("sobreprecio_leve")
    elif sobreprecio_pct <= -8:
        motivos.append("precio_bajo_mercado")

    if liquidez == "baja":
        motivos.append("mercado_liquidez_baja")
    elif liquidez == "media":
        motivos.append("mercado_liquidez_media")

    if competencia == "alto":
        motivos.append("alta_competencia")
    elif competencia == "medio":
        motivos.append("competencia_media")

    if brecha_pct <= -20:
        motivos.append("brecha_publicacion_cierre_muy_negativa")
    elif brecha_pct <= -10:
        motivos.append("brecha_publicacion_cierre_negativa")

    if score_presion >= 70:
        motivos.append("presion_comercial_muy_alta")
    elif score_presion >= 45:
        motivos.append("presion_comercial_alta")

    if estado_precio_tasacion in {"sobrevalorada", "sobreprecio_leve", "revision_dato"}:
        motivos.append(f"estado_precio_tasacion_{estado_precio_tasacion}")

    ratio_activo = (publicaciones_activas / publicaciones_totales) if publicaciones_totales > 0 else 0.0
    if publicaciones_activas >= 800 or ratio_activo >= 0.12:
        motivos.append("sobreoferta_publicaciones")

    # Dedupe conservando orden.
    seen = set()
    out: List[str] = []
    for m in motivos:
        if m not in seen:
            seen.add(m)
            out.append(m)
    return out


def build_argumento(
    codigo_propiedad: str,
    comuna: str,
    tipo_propiedad: str,
    precio_publicado_uf: float,
    tasacion_disponible: bool,
    tasacion_venta_uf: float,
    sobreprecio_pct: float,
    mercado: Dict[str, Any],
    campana: str,
    fuente_valorizacion: str,
    cantidad_leads_90d: int,
    dias_sin_leads: int,
    alto_interes_baja_conversion: bool,
    propiedad_invisible: bool,
) -> str:
    liquidez = mercado.get("liquidez", "media")
    competencia = mercado.get("nivel_competencia", "medio")
    brecha = safe_float(mercado.get("brecha_publicacion_vs_cierre_pct"))
    presion = safe_float(mercado.get("score_presion_comercial"))

    if tasacion_disponible:
        base = (
            f"Propiedad {codigo_propiedad} ({tipo_propiedad} en {comuna}) con precio publicado "
            f"{precio_publicado_uf:,.0f} UF y tasacion individual {tasacion_venta_uf:,.0f} UF "
            f"(sobreprecio {sobreprecio_pct:.1f}%)."
        )
    else:
        base = (
            f"Propiedad {codigo_propiedad} ({tipo_propiedad} en {comuna}) sin tasacion individual. "
            f"Se evalua con señales de mercado comunal ({fuente_valorizacion})."
        )

    contexto = (
        f" La comuna presenta liquidez {liquidez}, competencia {competencia}, "
        f"brecha publicacion/cierre {brecha:.1f}% y score de presion comercial {presion:.1f}."
    )

    if alto_interes_baja_conversion:
        leads_txt = (
            f" La propiedad ha generado interés comercial relevante ({cantidad_leads_90d} leads en 90d), "
            "pero el comportamiento sugiere fricción de precio frente al mercado."
        )
    elif propiedad_invisible:
        ds = "sin historial reciente de leads" if dias_sin_leads is None else f"{dias_sin_leads} días sin leads"
        leads_txt = (
            f" La propiedad presenta baja interacción comercial reciente (0 leads en 90d, {ds}), "
            "lo que indica baja tracción comercial."
        )
    elif campana == "destacar_oportunidad":
        leads_txt = " La propiedad mantiene un posicionamiento competitivo respecto al mercado comunal."
    else:
        leads_txt = f" Registra {cantidad_leads_90d} leads en 90d y {dias_sin_leads} días sin leads."

    if campana == "baja_precio_urgente":
        cierre = " Se recomienda ajuste de precio prioritario para reducir riesgo de estancamiento comercial."
    elif campana == "baja_precio_moderada":
        cierre = " Se recomienda ajuste moderado y seguimiento cercano de conversion."
    elif campana == "destacar_oportunidad":
        cierre = " Precio competitivo frente al mercado; conviene reforzar visibilidad y conversion."
    elif campana == "revision_datos":
        cierre = " Se requiere validacion de datos clave antes de activar gestion comercial."
    elif campana == "mantener_precio":
        cierre = " Precio alineado; mantener estrategia y monitoreo comercial."
    else:
        cierre = " Mantener seguimiento comercial con monitoreo semanal."

    return (base + contexto + leads_txt + cierre)[:900]


def compute_score(
    sobreprecio_pct: float,
    liquidez: str,
    competencia: str,
    brecha_pct: float,
    score_presion: float,
    estado_precio_tasacion: str,
    publicaciones_activas: int,
    publicaciones_totales: int,
) -> float:
    score = 0.0

    # 1) Sobreprecio
    if sobreprecio_pct > 20:
        score += 35
    elif sobreprecio_pct > 10:
        score += 25
    elif sobreprecio_pct > 5:
        score += 10

    # 2) Liquidez
    if liquidez == "baja":
        score += 15
    elif liquidez == "media":
        score += 7

    # 3) Competencia
    if competencia == "alto":
        score += 15
    elif competencia == "medio":
        score += 7

    # 4) Brecha publicacion/cierre
    if brecha_pct < -20:
        score += 20
    elif brecha_pct < -10:
        score += 10

    # 5) Score presion comercial
    score += score_presion * 0.35

    # 6) Estado precio desde tasacion
    if estado_precio_tasacion == "sobrevalorada":
        score += 12
    elif estado_precio_tasacion == "sobreprecio_leve":
        score += 6
    elif estado_precio_tasacion == "bajo_mercado":
        score -= 6

    # 7) Sobreoferta publicaciones
    ratio_activas = (publicaciones_activas / publicaciones_totales) if publicaciones_totales > 0 else 0.0
    if publicaciones_activas >= 800 or ratio_activas >= 0.12:
        score += 10
    elif publicaciones_activas >= 400 or ratio_activas >= 0.09:
        score += 5

    return clamp_score(score)


def choose_campaign(
    score_comercial: float,
    sobreprecio_pct: float,
    liquidez: str,
    competencia: str,
    brecha_pct: float,
    revision_datos_flag: bool,
) -> str:
    if revision_datos_flag:
        return "revision_datos"

    if sobreprecio_pct <= -8:
        return "destacar_oportunidad"

    strong_combo = liquidez == "baja" and competencia == "alto" and brecha_pct < -20

    if score_comercial >= 85 or sobreprecio_pct >= 20 or strong_combo:
        return "baja_precio_urgente"

    if 65 <= score_comercial <= 84:
        return "baja_precio_moderada"

    if score_comercial < 45:
        return "mantener_precio"

    return "seguimiento"


def score_interes_comercial(
    leads_90d: int,
    publicaciones_activas: int,
    nivel_competencia: str,
    liquidez: str,
) -> float:
    if leads_90d <= 0:
        base = 5.0
    elif leads_90d <= 2:
        base = 20.0
    elif leads_90d <= 5:
        base = 40.0
    elif leads_90d <= 10:
        base = 65.0
    else:
        base = 82.0

    if publicaciones_activas >= 800:
        base += 10
    elif publicaciones_activas >= 400:
        base += 6

    if nivel_competencia == "alto":
        base += 6
    elif nivel_competencia == "medio":
        base += 3

    if liquidez == "baja":
        base += 4
    elif liquidez == "media":
        base += 2
    return clamp_score(base)


def _autofit_ws(ws) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)


def compute_rebaja_sugerida(
    liquidez: str,
    sobreprecio_pct: float,
    precio_publicado_uf: float,
    tasacion_referencia_uf: float,
) -> Tuple[float, float, float]:
    if precio_publicado_uf <= 0 or tasacion_referencia_uf <= 0:
        return 0.0, 0.0, 0.0

    if liquidez == "alta":
        pct = 2.0
    elif liquidez == "media":
        pct = 4.0
    else:
        pct = 6.0

    if sobreprecio_pct > 20:
        pct = max(pct, 8.0)
    elif sobreprecio_pct > 10:
        pct = max(pct, 5.0)

    rebaja_uf = round(precio_publicado_uf * (pct / 100.0), 2)
    nuevo = round(max(precio_publicado_uf - rebaja_uf, tasacion_referencia_uf * 0.92), 2)
    return round(pct, 2), rebaja_uf, nuevo


def export_csv(col: Collection, export_path: Path, oficina_objetivo: str) -> int:
    export_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "codigo_propiedad",
        "operacion",
        "ejecutivo",
        "comuna",
        "tipo_propiedad",
        "precio_publicado_uf",
        "tasacion_venta_uf",
        "sobreprecio_pct",
        "score_comercial",
        "score_interes_comercial",
        "cantidad_leads_90d",
        "cantidad_leads_30d",
        "dias_sin_leads",
        "alto_interes_baja_conversion",
        "propiedad_invisible",
        "cluster_comercial",
        "accion_recomendada",
        "rebaja_sugerida_pct",
        "rebaja_sugerida_uf",
        "nuevo_precio_objetivo_uf",
        "riesgo_comercial",
        "campana_recomendada",
        "fuente_valorizacion",
        "tasacion_fuente_campo",
        "motivos_campana_txt",
        "resumen_mercado_comunal",
        "argumento_comercial",
    ]

    rows = list(
        col.find(
            {"oficina": oficina_objetivo},
            {
                "_id": 0,
                "codigo_propiedad": 1,
                "operacion": 1,
                "ejecutivo": 1,
                "comuna": 1,
                "tipo_propiedad": 1,
                "precio_publicado_uf": 1,
                "tasacion_venta_uf": 1,
                "sobreprecio_pct": 1,
                "score_comercial": 1,
                "score_interes_comercial": 1,
                "cantidad_leads_90d": 1,
                "cantidad_leads_30d": 1,
                "dias_sin_leads": 1,
                "alto_interes_baja_conversion": 1,
                "propiedad_invisible": 1,
                "cluster_comercial": 1,
                "accion_recomendada": 1,
                "rebaja_sugerida_pct": 1,
                "rebaja_sugerida_uf": 1,
                "nuevo_precio_objetivo_uf": 1,
                "riesgo_comercial": 1,
                "campana_recomendada": 1,
                "fuente_valorizacion": 1,
                "tasacion_fuente_campo": 1,
                "motivos_campana_txt": 1,
                "resumen_mercado_comunal": 1,
                "argumento_comercial": 1,
            },
        ).sort([("score_comercial", DESCENDING), ("codigo_propiedad", ASCENDING)])
    )

    with export_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in fields})
    return len(rows)


def export_xlsx(col: Collection, export_path: Path, oficina_objetivo: str) -> int:
    export_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "codigo_propiedad",
        "operacion",
        "ejecutivo",
        "comuna",
        "tipo_propiedad",
        "precio_publicado_uf",
        "tasacion_venta_uf",
        "sobreprecio_pct",
        "score_comercial",
        "score_interes_comercial",
        "cantidad_leads_90d",
        "cantidad_leads_30d",
        "dias_sin_leads",
        "alto_interes_baja_conversion",
        "propiedad_invisible",
        "cluster_comercial",
        "accion_recomendada",
        "rebaja_sugerida_pct",
        "rebaja_sugerida_uf",
        "nuevo_precio_objetivo_uf",
        "riesgo_comercial",
        "campana_recomendada",
        "fuente_valorizacion",
        "tasacion_fuente_campo",
        "motivos_campana_txt",
        "resumen_mercado_comunal",
        "argumento_comercial",
    ]
    rows = list(
        col.find(
            {"oficina": oficina_objetivo},
            {k: 1 for k in fields} | {"_id": 0},
        ).sort([("score_comercial", DESCENDING), ("codigo_propiedad", ASCENDING)])
    )
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.reindex(columns=fields)
    else:
        df = pd.DataFrame(columns=fields)

    detail_name = "Detalle"
    resumen_name = "Resumen_Ejecutivo"
    try:
        with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=detail_name, index=False)

            wb = writer.book
            ws = wb[detail_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            _autofit_ws(ws)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="D9E1F2")
            arg_col = None
            for idx, c in enumerate(ws[1], start=1):
                if c.value == "argumento_comercial":
                    arg_col = idx
                    break
            if arg_col:
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=arg_col).alignment = Alignment(wrap_text=True, vertical="top")
                ws.column_dimensions[get_column_letter(arg_col)].width = 55

            # Hoja ejecutiva
            rs = wb.create_sheet(resumen_name)
            bold = Font(bold=True)
            title_font = Font(bold=True, size=14)
            head_fill = PatternFill("solid", fgColor="BDD7EE")

            total_prop = len(df)
            urg = int((df["accion_recomendada"] == "bajar_precio_urgente").sum()) if total_prop else 0
            sug = int((df["accion_recomendada"] == "bajar_precio_sugerida").sum()) if total_prop else 0
            rev = int((df["accion_recomendada"] == "revisar_publicacion").sum()) if total_prop else 0
            dest = int((df["accion_recomendada"] == "destacar_propiedad").sum()) if total_prop else 0
            val = int((df["accion_recomendada"] == "validar_datos").sum()) if total_prop else 0
            sobreval = int((df["sobreprecio_pct"] > 0).sum()) if total_prop else 0
            pct_sobre = (sobreval / total_prop) if total_prop else 0.0
            sobre_prom = float(df["sobreprecio_pct"].mean()) if total_prop else 0.0
            rebaja_total = float(df["rebaja_sugerida_uf"].fillna(0).sum()) if total_prop else 0.0
            rebaja_pct_prom = float(df["rebaja_sugerida_pct"].fillna(0).mean()) if total_prop else 0.0

            r = 1
            rs.cell(r, 1, "Resumen Ejecutivo Comercial").font = title_font
            r += 2
            rs.cell(r, 1, "KPI Generales").font = bold
            r += 1
            kpis = [
                ("total_propiedades", total_prop),
                ("total_bajar_precio_urgente", urg),
                ("total_bajar_precio_sugerida", sug),
                ("total_revisar_publicacion", rev),
                ("total_destacar_propiedad", dest),
                ("total_validar_datos", val),
                ("porcentaje_sobrevaloradas", pct_sobre),
                ("sobreprecio_promedio_pct", sobre_prom / 100.0),
                ("rebaja_total_sugerida_uf", rebaja_total),
                ("rebaja_promedio_sugerida_pct", rebaja_pct_prom / 100.0),
            ]
            for k, v in kpis:
                rs.cell(r, 1, k)
                rs.cell(r, 2, v)
                r += 1
            r += 1

            def write_table(start_row: int, title: str, table_df: pd.DataFrame) -> int:
                rs.cell(start_row, 1, title).font = bold
                start_row += 1
                for j, coln in enumerate(table_df.columns, start=1):
                    c = rs.cell(start_row, j, coln)
                    c.font = bold
                    c.fill = head_fill
                start_row += 1
                for _, rowv in table_df.iterrows():
                    for j, coln in enumerate(table_df.columns, start=1):
                        rs.cell(start_row, j, rowv[coln])
                    start_row += 1
                return start_row + 1

            # Resumen por accion
            if total_prop:
                t_acc = (
                    df.groupby("accion_recomendada", dropna=False).size().reset_index(name="cantidad")
                )
                t_acc["porcentaje"] = t_acc["cantidad"] / total_prop
                t_acc = t_acc.sort_values("cantidad", ascending=False)
            else:
                t_acc = pd.DataFrame(columns=["accion_recomendada", "cantidad", "porcentaje"])
            r = write_table(r, "Resumen por accion_recomendada", t_acc)

            # Resumen por cluster
            if total_prop:
                t_cluster = df.groupby("cluster_comercial", dropna=False).size().reset_index(name="cantidad").sort_values("cantidad", ascending=False)
            else:
                t_cluster = pd.DataFrame(columns=["cluster_comercial", "cantidad"])
            r = write_table(r, "Resumen por cluster_comercial", t_cluster)

            # Resumen por comuna (solo campañas de baja)
            df_baja = df[df["accion_recomendada"].isin(["bajar_precio_urgente", "bajar_precio_sugerida"])].copy()
            if not df_baja.empty:
                grp = df_baja.groupby("comuna", dropna=False).agg(
                    propiedades=("codigo_propiedad", "count"),
                    promedio_sobreprecio_pct=("sobreprecio_pct", "mean"),
                    leads_promedio_90d=("cantidad_leads_90d", "mean"),
                    rebaja_total_uf=("rebaja_sugerida_uf", "sum"),
                    urgentes=("accion_recomendada", lambda s: int((s == "bajar_precio_urgente").sum())),
                ).reset_index()
                grp = grp.sort_values(["urgentes", "promedio_sobreprecio_pct"], ascending=[False, False]).drop(columns=["urgentes"])
            else:
                grp = pd.DataFrame(columns=["comuna", "propiedades", "promedio_sobreprecio_pct", "leads_promedio_90d", "rebaja_total_uf"])
            r = write_table(r, "Resumen por comuna (campanas de baja)", grp)

            # Tasacion vs publicado
            df_valid = df[(df["tasacion_venta_uf"] > 0) & (df["precio_publicado_uf"] > 0)].copy()
            if not df_valid.empty:
                dif_uf = df_valid["precio_publicado_uf"] - df_valid["tasacion_venta_uf"]
                comp = pd.DataFrame(
                    [
                        ["cantidad_propiedades_sobre_tasacion", int((dif_uf > 0).sum())],
                        ["cantidad_propiedades_bajo_tasacion", int((dif_uf < 0).sum())],
                        ["cantidad_propiedades_rango_competitivo", int((df_valid["sobreprecio_pct"].abs() <= 5).sum())],
                        ["diferencia_promedio_uf", float(dif_uf.mean())],
                        ["diferencia_promedio_pct", float(df_valid["sobreprecio_pct"].mean()) / 100.0],
                    ],
                    columns=["metrica", "valor"],
                )
            else:
                comp = pd.DataFrame(columns=["metrica", "valor"])
            r = write_table(r, "Resumen tasacion vs precio publicado", comp)

            # Resumen leads
            leads_altos = int((df["cantidad_leads_90d"] >= 8).sum()) if total_prop else 0
            leads0 = int((df["cantidad_leads_90d"] == 0).sum()) if total_prop else 0
            prom90 = float(df["cantidad_leads_90d"].mean()) if total_prop else 0.0
            m_leads = pd.DataFrame(
                [
                    ["propiedades_con_0_leads", leads0],
                    ["propiedades_con_leads_altos", leads_altos],
                    ["promedio_leads_90d", prom90],
                ],
                columns=["metrica", "valor"],
            )
            r = write_table(r, "Resumen leads", m_leads)

            top10_leads = df.sort_values(["cantidad_leads_90d", "score_comercial"], ascending=[False, False]).head(10)[
                ["codigo_propiedad", "comuna", "cantidad_leads_90d", "sobreprecio_pct", "accion_recomendada"]
            ]
            r = write_table(r, "Top 10 propiedades con mas leads", top10_leads)

            top20 = df[df["accion_recomendada"] == "bajar_precio_urgente"].sort_values(
                ["score_comercial", "sobreprecio_pct"], ascending=[False, False]
            ).head(20)[
                [
                    "codigo_propiedad",
                    "comuna",
                    "precio_publicado_uf",
                    "tasacion_venta_uf",
                    "sobreprecio_pct",
                    "cantidad_leads_90d",
                    "rebaja_sugerida_pct",
                    "rebaja_sugerida_uf",
                    "cluster_comercial",
                    "argumento_comercial",
                ]
            ]
            top20 = top20.rename(columns={"tasacion_venta_uf": "tasacion_referencia_uf"})
            r = write_table(r, "Top 20 propiedades prioritarias", top20)

            rs.freeze_panes = "A2"
            rs.auto_filter.ref = rs.dimensions
            _autofit_ws(rs)
            # formatos
            for row in rs.iter_rows(min_row=1, max_row=rs.max_row, min_col=1, max_col=rs.max_column):
                for c in row:
                    if isinstance(c.value, float):
                        if "pct" in str(rs.cell(row=1, column=c.column).value).lower():
                            c.number_format = "0.00%"
            for c in rs["A"]:
                if c.value == "argumento_comercial":
                    col = c.column
                    for rr in range(c.row + 1, rs.max_row + 1):
                        rs.cell(rr, col).alignment = Alignment(wrap_text=True, vertical="top")
                    rs.column_dimensions[get_column_letter(col)].width = 65
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = export_path.with_name(f"{export_path.stem}_{ts}{export_path.suffix}")
        with pd.ExcelWriter(fallback, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=detail_name, index=False)
        logger.warning("XLSX bloqueado. Export alternativo generado en %s", fallback)
    return len(rows)


def extract_codigo_from_lead(doc: Dict[str, Any]) -> str:
    candidates = [
        doc.get("codigo_propiedad"),
        doc.get("codigo"),
        doc.get("id_propiedad"),
        (doc.get("prospecto") or {}).get("codigo"),
        (doc.get("prospecto") or {}).get("codigo_propiedad"),
        (doc.get("prospecto") or {}).get("id_propiedad"),
    ]
    for c in candidates:
        s = str(c or "").strip()
        if s:
            return s
    return ""


def extract_lead_datetime(doc: Dict[str, Any]) -> Optional[datetime]:
    dates = [
        parse_dt(doc.get("created_at")),
        parse_dt(doc.get("ultima_actualizacion_bi")),
        parse_dt((doc.get("prospecto") or {}).get("ultimo_mensaje")),
    ]
    msgs = doc.get("messages") or []
    if isinstance(msgs, list):
        for m in reversed(msgs):
            dt = parse_dt((m or {}).get("timestamp"))
            if dt:
                dates.append(dt)
                break
    dates = [d for d in dates if d]
    if not dates:
        return None
    return max(dates)


def build_leads_cache(leads_col: Collection) -> Dict[str, Dict[str, Any]]:
    now = datetime.now(timezone.utc)
    since_90 = now - timedelta(days=90)
    since_30 = now - timedelta(days=30)
    out: Dict[str, Dict[str, Any]] = {}
    for d in leads_col.find({}, {"_id": 0, "codigo_propiedad": 1, "codigo": 1, "id_propiedad": 1, "prospecto.codigo": 1, "prospecto.codigo_propiedad": 1, "prospecto.id_propiedad": 1, "created_at": 1, "ultima_actualizacion_bi": 1, "prospecto.ultimo_mensaje": 1, "messages.timestamp": 1}):
        codigo = extract_codigo_from_lead(d)
        if not codigo:
            continue
        dt = extract_lead_datetime(d)
        if not dt:
            continue
        agg = out.setdefault(codigo, {"cantidad_leads_90d": 0, "cantidad_leads_30d": 0, "fecha_ultimo_lead": None})
        if dt >= since_90:
            agg["cantidad_leads_90d"] += 1
        if dt >= since_30:
            agg["cantidad_leads_30d"] += 1
        curr_last = parse_dt(agg.get("fecha_ultimo_lead"))
        if not curr_last or dt > curr_last:
            agg["fecha_ultimo_lead"] = dt.isoformat()

    for codigo, agg in out.items():
        last_dt = parse_dt(agg.get("fecha_ultimo_lead"))
        if last_dt:
            agg["dias_sin_leads"] = max(0, (now - last_dt).days)
        else:
            agg["dias_sin_leads"] = None
    return out


def build_doc(
    cartera: Dict[str, Any],
    tas_doc: Optional[Dict[str, Any]],
    mercado_doc: Optional[Dict[str, Any]],
    leads_stats: Dict[str, Any],
) -> Dict[str, Any]:
    codigo = str(cartera.get("codigo") or "").strip()
    precio_publicado_uf = safe_float(cartera.get("precio_uf"))
    ejecutivo = str(cartera.get("ejecutivo") or cartera.get("broker") or "")
    operacion = str(cartera.get("operacion") or "venta").strip().lower()

    comuna_raw = str(cartera.get("comuna") or "").strip()
    tipo_raw = str(cartera.get("tipo") or "").strip()
    tipo_base = normalize_tipo_propiedad(tipo_raw)
    comuna, tipo_propiedad, match_key = normalize_comuna_tipo(comuna_raw, tipo_base)

    tasacion_disponible = False
    tasacion_venta_uf = 0.0
    tasacion_fuente_campo = "sin_tasacion"
    sobreprecio_pct = 0.0
    fuente_valorizacion = "mercado_comunal"
    estado_precio_tasacion = ""

    if tas_doc:
        tasacion_venta_uf, tasacion_fuente_campo = extract_tasacion_uf(tas_doc, operacion=operacion)
        if tasacion_venta_uf > 0:
            tasacion_disponible = True
            fuente_valorizacion = "tasacion_individual"
            if precio_publicado_uf > 0:
                sobreprecio_pct = round(((precio_publicado_uf - tasacion_venta_uf) / tasacion_venta_uf) * 100.0, 2)
        estado_precio_tasacion = str(((tas_doc.get("analisis_comercial") or {}).get("estado_precio") or "")).strip()

    indicadores = ((mercado_doc or {}).get("indicadores_mercado") or {})
    mercado_venta = ((mercado_doc or {}).get("mercado_venta") or {})
    resumen_mercado_comunal = str((mercado_doc or {}).get("resumen_comercial_llm") or "")
    cantidad_leads_90d = int(leads_stats.get("cantidad_leads_90d") or 0)
    cantidad_leads_30d = int(leads_stats.get("cantidad_leads_30d") or 0)
    fecha_ultimo_lead = leads_stats.get("fecha_ultimo_lead")
    dias_sin_leads = leads_stats.get("dias_sin_leads")

    liquidez = str(indicadores.get("liquidez") or "media")
    presion_baja_precio = str(indicadores.get("presion_baja_precio") or "media")
    nivel_competencia = str(indicadores.get("nivel_competencia") or "medio")
    score_presion = safe_float(indicadores.get("score_presion_comercial"))
    brecha_pct = safe_float(indicadores.get("brecha_publicacion_vs_cierre_pct"))
    publicaciones_activas = int(safe_float(mercado_venta.get("publicaciones_activas")))
    publicaciones_totales = int(safe_float(mercado_venta.get("publicaciones_totales")))

    # Si no hay tasacion individual, intentamos inferir una referencia por UF/m2 comunal cuando existan m2.
    if not tasacion_disponible:
        m2_construidos = safe_float(
            cartera.get("metros_construidos")
            or cartera.get("m2")
            or cartera.get("superficie_m2")
            or cartera.get("superficie_util")
            or cartera.get("sup_util")
        )
        if operacion == "arriendo":
            uf_m2_efectiva = safe_float(((mercado_doc or {}).get("mercado_arriendo") or {}).get("uf_m2_arriendo_actual"))
        else:
            uf_m2_efectiva = safe_float(mercado_venta.get("uf_m2_venta_efectiva_actual"))
        tas_ref, metodo_ref = infer_tasacion_uf_from_comuna(precio_publicado_uf, m2_construidos, uf_m2_efectiva)
        if tas_ref > 0:
            tasacion_venta_uf = tas_ref
            sobreprecio_pct = round(((precio_publicado_uf - tasacion_venta_uf) / tasacion_venta_uf) * 100.0, 2) if tasacion_venta_uf > 0 else 0.0
            fuente_valorizacion = f"mercado_comunal_{metodo_ref}"
        else:
            tasacion_venta_uf = 0.0
            sobreprecio_pct = 0.0
            fuente_valorizacion = "mercado_comunal"

    revision_datos_flag = is_revision_datos(comuna, tipo_propiedad, precio_publicado_uf)

    score_comercial = compute_score(
        sobreprecio_pct=sobreprecio_pct,
        liquidez=liquidez,
        competencia=nivel_competencia,
        brecha_pct=brecha_pct,
        score_presion=score_presion,
        estado_precio_tasacion=estado_precio_tasacion,
        publicaciones_activas=publicaciones_activas,
        publicaciones_totales=publicaciones_totales,
    )

    if revision_datos_flag:
        score_comercial = max(score_comercial, 55.0)

    score_interes = score_interes_comercial(
        leads_90d=cantidad_leads_90d,
        publicaciones_activas=publicaciones_activas,
        nivel_competencia=nivel_competencia,
        liquidez=liquidez,
    )

    alto_interes_baja_conversion = (
        cantidad_leads_90d >= 6
        and sobreprecio_pct > 8
        and liquidez in {"media", "baja"}
    )
    propiedad_invisible = (
        cantidad_leads_90d == 0
        and score_comercial >= 70
        and nivel_competencia in {"alto", "medio"}
    )

    campana = choose_campaign(
        score_comercial=score_comercial,
        sobreprecio_pct=sobreprecio_pct,
        liquidez=liquidez,
        competencia=nivel_competencia,
        brecha_pct=brecha_pct,
        revision_datos_flag=revision_datos_flag,
    )

    if cantidad_leads_90d >= 8 and sobreprecio_pct > 8:
        cluster_comercial = "alto_interes_precio_alto"
    elif cantidad_leads_90d == 0:
        cluster_comercial = "invisible"
    elif cantidad_leads_90d <= 2 and nivel_competencia == "alto" and liquidez == "baja":
        cluster_comercial = "mercado_saturado"
    elif cantidad_leads_90d >= 4 and sobreprecio_pct <= 5:
        cluster_comercial = "oportunidad"
    else:
        cluster_comercial = "estable"

    anti_error = (
        tasacion_venta_uf <= 0
        or precio_publicado_uf <= 0
        or (not comuna.strip()) or comuna.strip().lower() == "desconocido"
        or abs(sobreprecio_pct) > 80
    )

    if anti_error:
        accion_recomendada = "validar_datos"
    elif score_comercial >= 85 and sobreprecio_pct >= 12 and (cantidad_leads_90d <= 2 or alto_interes_baja_conversion):
        accion_recomendada = "bajar_precio_urgente"
    elif alto_interes_baja_conversion or (score_comercial >= 65 and sobreprecio_pct > 8):
        accion_recomendada = "bajar_precio_sugerida"
    elif propiedad_invisible and sobreprecio_pct <= 8:
        accion_recomendada = "revisar_publicacion"
    elif cluster_comercial == "oportunidad":
        accion_recomendada = "destacar_propiedad"
    else:
        accion_recomendada = "mantener_precio"

    rebaja_pct, rebaja_uf, nuevo_precio = compute_rebaja_sugerida(
        liquidez=liquidez,
        sobreprecio_pct=sobreprecio_pct,
        precio_publicado_uf=precio_publicado_uf,
        tasacion_referencia_uf=tasacion_venta_uf,
    )
    if accion_recomendada not in {"bajar_precio_urgente", "bajar_precio_sugerida"}:
        rebaja_pct, rebaja_uf, nuevo_precio = 0.0, 0.0, 0.0

    motivos = build_motivos(
        sobreprecio_pct=sobreprecio_pct,
        liquidez=liquidez,
        competencia=nivel_competencia,
        brecha_pct=brecha_pct,
        score_presion=score_presion,
        estado_precio_tasacion=estado_precio_tasacion,
        publicaciones_activas=publicaciones_activas,
        publicaciones_totales=publicaciones_totales,
        revision_datos_flag=revision_datos_flag,
    )

    mercado_obj = {
        "liquidez": liquidez,
        "presion_baja_precio": presion_baja_precio,
        "nivel_competencia": nivel_competencia,
        "score_presion_comercial": round(score_presion, 2),
        "brecha_publicacion_vs_cierre_pct": round(brecha_pct, 2),
    }

    argumento = build_argumento(
        codigo_propiedad=codigo,
        comuna=comuna,
        tipo_propiedad=tipo_propiedad,
        precio_publicado_uf=precio_publicado_uf,
        tasacion_disponible=tasacion_disponible,
        tasacion_venta_uf=tasacion_venta_uf,
        sobreprecio_pct=sobreprecio_pct,
        mercado=mercado_obj,
        campana=campana,
        fuente_valorizacion=fuente_valorizacion,
        cantidad_leads_90d=cantidad_leads_90d,
        dias_sin_leads=dias_sin_leads,
        alto_interes_baja_conversion=alto_interes_baja_conversion,
        propiedad_invisible=propiedad_invisible,
    )

    return {
        "codigo_propiedad": codigo,
        "ejecutivo": ejecutivo,
        "oficina": str(cartera.get("oficina") or ""),
        "operacion": operacion,
        "comuna": comuna,
        "tipo_propiedad": tipo_propiedad,
        "precio_publicado_uf": round(precio_publicado_uf, 2),
        "tasacion_disponible": tasacion_disponible,
        "tasacion_venta_uf": round(tasacion_venta_uf, 2),
        "sobreprecio_pct": round(sobreprecio_pct, 2),
        "mercado": mercado_obj,
        "score_comercial": round(score_comercial, 2),
        "score_interes_comercial": round(score_interes, 2),
        "cantidad_leads_90d": cantidad_leads_90d,
        "cantidad_leads_30d": cantidad_leads_30d,
        "fecha_ultimo_lead": fecha_ultimo_lead,
        "dias_sin_leads": dias_sin_leads,
        "alto_interes_baja_conversion": alto_interes_baja_conversion,
        "propiedad_invisible": propiedad_invisible,
        "cluster_comercial": cluster_comercial,
        "accion_recomendada": accion_recomendada,
        "rebaja_sugerida_pct": rebaja_pct,
        "rebaja_sugerida_uf": rebaja_uf,
        "nuevo_precio_objetivo_uf": nuevo_precio,
        "riesgo_comercial": risk_level(score_comercial),
        "campana_recomendada": campana,
        "motivos_campana": motivos,
        "motivos_campana_txt": " | ".join(motivos),
        "argumento_comercial": argumento,
        "resumen_mercado_comunal": resumen_mercado_comunal[:800],
        "fuente_valorizacion": fuente_valorizacion,
        "tasacion_fuente_campo": tasacion_fuente_campo,
        "ready_para_campana": not revision_datos_flag,
        "insumos_recomendacion": {
            "operacion": operacion,
            "precio_publicado_uf": round(precio_publicado_uf, 2),
            "tasacion_venta_uf": round(tasacion_venta_uf, 2),
            "sobreprecio_pct": round(sobreprecio_pct, 2),
            "liquidez": liquidez,
            "competencia": nivel_competencia,
            "brecha_publicacion_vs_cierre_pct": round(brecha_pct, 2),
            "score_presion_comercial": round(score_presion, 2),
            "estado_precio_tasacion": estado_precio_tasacion,
            "publicaciones_activas": publicaciones_activas,
            "publicaciones_totales": publicaciones_totales,
        },
        "qa_flags": {
            "sin_match_mercado_comunal": mercado_doc is None,
            "sin_tasacion_individual": not tasacion_disponible,
            "revision_datos": revision_datos_flag,
        },
        "match_key": match_key,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


def run_pipeline(
    limit: Optional[int] = None,
    batch_size: int = 500,
    include_non_disponible: bool = False,
    oficina_objetivo: str = "PROCASA SUCRE",
) -> Dict[str, int]:
    db = get_db()
    col_tas = db["tasaciones"]
    col_mc = db["mercado_comunal"]
    col_uc = db["universo_cartera"]
    col_leads = db["leads"]
    col_out = db["propiedades_accionables"]

    ensure_indexes(col_out)
    col_out.update_many(
        {"oficina": oficina_objetivo},
        {"$unset": {"precio_publicado_uf_redondeado": "", "tasacion_referencia_uf_redondeada": ""}},
    )

    # Cache tasaciones por codigo
    tas_cache: Dict[str, Dict[str, Any]] = {}
    for t in col_tas.find(
        {"codigo_propiedad": {"$exists": True, "$ne": ""}},
        {
            "codigo_propiedad": 1,
            "tasacion_online.valor_comercial.uf": 1,
            "tasacion_online.arriendo_estimado.uf": 1,
            "tasacion_online.valor_minimo_maximo.estimacion_valor_comercial_uf": 1,
            "tasacion_online.valor_minimo_maximo.precio_maximo_uf": 1,
            "tasacion_online.valor_minimo_maximo.precio_minimo_uf": 1,
            "analisis_comercial.estado_precio": 1,
        },
    ):
        codigo = str(t.get("codigo_propiedad") or "").strip()
        if codigo:
            tas_cache[codigo] = t

    # Cache mercado comunal por match key.
    mercado_cache: Dict[str, Dict[str, Any]] = {}
    for m in col_mc.find(
        {},
        {
            "comuna": 1,
            "tipo_propiedad": 1,
            "indicadores_mercado": 1,
            "mercado_venta": 1,
            "mercado_arriendo": 1,
            "resumen_comercial_llm": 1,
        },
    ):
        mk = build_match_key(str(m.get("comuna") or ""), str(m.get("tipo_propiedad") or ""))
        if mk:
            mercado_cache[mk] = m

    leads_cache = build_leads_cache(col_leads)

    cartera_query: Dict[str, Any]
    if include_non_disponible:
        cartera_query = {
            "codigo": {"$exists": True, "$ne": ""},
            "oficina": oficina_objetivo,
        }
    else:
        cartera_query = {
            "codigo": {"$exists": True, "$ne": ""},
            "disponible": True,
            "oficina": oficina_objetivo,
        }

    total = col_uc.count_documents(cartera_query)
    cursor = col_uc.find(cartera_query)
    if limit and limit > 0:
        cursor = cursor.limit(limit)

    ops: List[UpdateOne] = []

    stats = {
        "procesadas": 0,
        "sin_tasacion": 0,
        "sin_match_mercado": 0,
        "ready_para_campana": 0,
        "revision_datos": 0,
    }

    progress_total = min(total, limit) if (limit and limit > 0) else total

    for cartera in tqdm(cursor, total=progress_total, desc="Propiedades accionables", unit="prop", dynamic_ncols=True):
        codigo = str(cartera.get("codigo") or "").strip()
        if not codigo:
            continue

        comuna_raw = str(cartera.get("comuna") or "")
        tipo_raw = str(cartera.get("tipo") or "")
        tipo_base = normalize_tipo_propiedad(tipo_raw)
        comuna_norm, tipo_norm, match_key = normalize_comuna_tipo(comuna_raw, tipo_base)

        tas_doc = tas_cache.get(codigo)
        mercado_doc = mercado_cache.get(match_key)

        leads_stats = leads_cache.get(codigo, {"cantidad_leads_90d": 0, "cantidad_leads_30d": 0, "fecha_ultimo_lead": None, "dias_sin_leads": None})
        out_doc = build_doc(cartera=cartera, tas_doc=tas_doc, mercado_doc=mercado_doc, leads_stats=leads_stats)

        if out_doc["qa_flags"]["sin_tasacion_individual"]:
            stats["sin_tasacion"] += 1
        if out_doc["qa_flags"]["sin_match_mercado_comunal"]:
            stats["sin_match_mercado"] += 1
        if out_doc["ready_para_campana"]:
            stats["ready_para_campana"] += 1
        if out_doc["qa_flags"]["revision_datos"]:
            stats["revision_datos"] += 1

        ops.append(
            UpdateOne(
                {"codigo_propiedad": codigo},
                {"$set": out_doc, "$setOnInsert": {"created_at": datetime.now(timezone.utc).isoformat()}},
                upsert=True,
            )
        )
        stats["procesadas"] += 1

        if len(ops) >= batch_size:
            col_out.bulk_write(ops, ordered=False)
            ops.clear()

    if ops:
        col_out.bulk_write(ops, ordered=False)

    n_export_xlsx = export_xlsx(
        col_out,
        PROJECT_ROOT / "exports" / "propiedades_accionables.xlsx",
        oficina_objetivo=oficina_objetivo,
    )
    logger.info("XLSX exportado con %s filas", n_export_xlsx)

    return stats


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Pipeline propiedades_accionables")
    ap.add_argument("--limit", type=int, default=0, help="Procesar solo N propiedades")
    ap.add_argument("--batch-size", type=int, default=500, help="Batch size para bulk_write")
    ap.add_argument("--include-non-disponible", action="store_true", help="Incluye propiedades no disponibles")
    ap.add_argument("--oficina", default="PROCASA SUCRE", help="Filtra universo_cartera por oficina")
    ap.add_argument("--log-level", default="INFO", help="DEBUG/INFO/WARNING/ERROR")
    return ap.parse_args()


def main() -> None:
    args = parse_args()
    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    )

    stats = run_pipeline(
        limit=args.limit or None,
        batch_size=args.batch_size,
        include_non_disponible=args.include_non_disponible,
        oficina_objetivo=args.oficina,
    )
    logger.info("Pipeline completado: %s", stats)


if __name__ == "__main__":
    main()
